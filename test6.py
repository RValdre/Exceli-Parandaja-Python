import openpyxl as xl
wbook = xl.load_workbook('a.xlsx')
sheet = wbook.active

print("Data validation'i näitemine:")
for data_val in sheet.data_validations.dataValidation:
    print("Veerg:", data_val)
    for data in data_val:
        print(data)

print("\nLahtrite formaadid:")
for row in sheet:
    for elem in row:
        print(elem.value, elem.number_format)

print("\nInfo fondi kohta:")
print(sheet['C7'].font)
print(sheet.max_row, sheet.max_column)

print("\nValemi tükeldamine:")
formula = sheet['C7'].value
tokenized_formula = xl.formula.Tokenizer(formula)
for token in tokenized_formula.items:
    print("%10s%10s%10s" % (token.value, token.type, token.subtype))

wbook.close()